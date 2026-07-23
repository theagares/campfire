"""
app/core/parser/xlsx.py
XLSX 파싱 — openpyxl (PLAN §6).

모든 시트를 순회하며 셀 텍스트를 합쳐 하나의 문서 텍스트로 반환한다
(txt.py/docx.py 와 동일하게 예외를 밖으로 던지고, 상위 parser/__init__.py 가
"미검사 통과" 정책(PLAN §9.2)에 맞춰 scan_status 로 변환한다).
"""

from __future__ import annotations

import io


def extract_xlsx(file_bytes: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        sheets: list[str] = []
        for ws in wb.worksheets:
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"[{ws.title}]\n" + "\n".join(rows))
        return "\n\n---\n\n".join(sheets)
    finally:
        wb.close()
